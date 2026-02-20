import logging

import httpx

from agentic_fs.config import settings

logger = logging.getLogger(__name__)


class ExtractedText:
    def __init__(self, text: str, method: str, char_count: int):
        self.text = text
        self.method = method
        self.char_count = char_count


class Extractor:
    def __init__(self):
        self.tika_url = settings.tika_url

    def extract(self, file_path: str, mime_type: str) -> ExtractedText:
        # For plain text types, just read directly
        if mime_type.startswith("text/") or mime_type in (
            "application/json",
            "application/xml",
            "application/javascript",
            "application/x-python",
        ):
            return self._read_text(file_path)

        # Try Tika first for binary types
        try:
            return self._extract_with_tika(file_path, mime_type)
        except Exception as e:
            logger.warning(f"Tika extraction failed: {e}, trying fallback")

        # Fallback to Python libraries
        return self._fallback_extract(file_path, mime_type)

    def _read_text(self, file_path: str) -> ExtractedText:
        with open(file_path, "r", errors="replace") as f:
            text = f.read()
        return ExtractedText(text=text, method="direct_read", char_count=len(text))

    def _extract_with_tika(self, file_path: str, mime_type: str) -> ExtractedText:
        with open(file_path, "rb") as f:
            content = f.read()

        response = httpx.put(
            f"{self.tika_url}/tika",
            content=content,
            headers={
                "Content-Type": mime_type,
                "Accept": "text/plain",
            },
            timeout=60.0,
        )
        response.raise_for_status()
        text = response.text.strip()

        if not text:
            raise ValueError("Tika returned empty text")

        return ExtractedText(text=text, method="tika", char_count=len(text))

    def _fallback_extract(self, file_path: str, mime_type: str) -> ExtractedText:
        if mime_type == "application/pdf":
            return self._extract_pdf(file_path)
        elif mime_type in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ):
            return self._extract_docx(file_path)
        elif mime_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            return self._extract_xlsx(file_path)
        elif mime_type.startswith("image/"):
            return self._extract_image(file_path)
        else:
            # Last resort: try reading as text
            try:
                return self._read_text(file_path)
            except Exception:
                return ExtractedText(text="", method="none", char_count=0)

    def _extract_pdf(self, file_path: str) -> ExtractedText:
        from pypdf import PdfReader

        reader = PdfReader(file_path)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        text = "\n\n".join(pages)
        return ExtractedText(text=text, method="pypdf", char_count=len(text))

    def _extract_docx(self, file_path: str) -> ExtractedText:
        from docx import Document

        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)
        return ExtractedText(text=text, method="python-docx", char_count=len(text))

    def _extract_xlsx(self, file_path: str) -> ExtractedText:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True)
        rows = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows.append(f"Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(",".join(cells))
        wb.close()
        text = "\n".join(rows)
        return ExtractedText(text=text, method="openpyxl", char_count=len(text))

    def _extract_image(self, file_path: str) -> ExtractedText:
        try:
            from PIL import Image
            import pytesseract

            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return ExtractedText(text=text, method="pytesseract", char_count=len(text))
        except Exception as e:
            logger.warning(f"OCR extraction failed: {e}")
            return ExtractedText(
                text=f"[Image file: {file_path}]", method="none", char_count=0
            )
